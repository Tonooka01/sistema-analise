"""
routes_dre.py
Blueprint para DRE — Demonstração do Resultado do Exercício.
"""

import io
import math
import sqlite3
from datetime import datetime as dt

from flask import Blueprint, jsonify, request, current_app
from flask_login import login_required

from logger import get_logger

logger = get_logger(__name__)

dre_bp = Blueprint('dre_bp', __name__)


def get_db():
    return current_app.config['GET_DB_CONNECTION']()


# ---------------------------------------------------------------------------
# GET /filters
# ---------------------------------------------------------------------------

@dre_bp.route('/filters')
@login_required
def api_dre_filters():
    conn = get_db()
    try:
        grupos    = [r[0] for r in conn.execute(
            "SELECT DISTINCT Grupo_DRE FROM DRE WHERE Grupo_DRE IS NOT NULL ORDER BY Grupo_DRE"
        ).fetchall()]
        subgrupos = [r[0] for r in conn.execute(
            "SELECT DISTINCT Subgrupo_DRE FROM DRE WHERE Subgrupo_DRE IS NOT NULL ORDER BY Subgrupo_DRE"
        ).fetchall()]
        situacoes = [r[0] for r in conn.execute(
            "SELECT DISTINCT Situacao FROM DRE WHERE Situacao IS NOT NULL ORDER BY Situacao"
        ).fetchall()]
        anos      = [r[0] for r in conn.execute(
            "SELECT DISTINCT Ano FROM DRE WHERE Ano IS NOT NULL ORDER BY Ano DESC"
        ).fetchall()]
        lojas     = [r[0] for r in conn.execute(
            "SELECT DISTINCT Loja FROM DRE WHERE Loja IS NOT NULL ORDER BY Loja"
        ).fetchall()]
        return jsonify({
            "grupos": grupos,
            "subgrupos": subgrupos,
            "situacoes": situacoes,
            "anos": anos,
            "lojas": lojas,
        })
    except sqlite3.Error as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /dre_report  — Receita (CAR) × Despesas (DRE) por período
# ---------------------------------------------------------------------------

@dre_bp.route('/dre_report')
@login_required
def api_dre_report():
    conn = get_db()
    try:
        start_date = request.args.get('start_date', '')
        end_date   = request.args.get('end_date', '')
        situacao   = request.args.get('situacao', '')

        # --- Receita: Contas_a_Receber ---
        car_c = ["CR.Status = 'Recebido'", "CR.Data_pagamento IS NOT NULL", "CR.Data_pagamento != ''"]
        car_p = []
        if start_date:
            car_c.append("CR.Data_pagamento >= ?"); car_p.append(start_date)
        if end_date:
            car_c.append("CR.Data_pagamento <= ?"); car_p.append(end_date)

        receita_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', CR.Data_pagamento) AS periodo,
                   SUM(CR.Valor_recebido) AS valor
            FROM Contas_a_Receber CR
            WHERE {' AND '.join(car_c)}
            GROUP BY periodo ORDER BY periodo
        """, car_p).fetchall()

        # --- Despesas: DRE ---
        dre_c = ["Data_Competencia IS NOT NULL", "Data_Competencia != ''"]
        dre_p = []
        if start_date:
            dre_c.append("Data_Competencia >= ?"); dre_p.append(start_date)
        if end_date:
            dre_c.append("Data_Competencia <= ?"); dre_p.append(end_date)
        if situacao:
            dre_c.append("Situacao = ?"); dre_p.append(situacao)

        dre_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', Data_Competencia) AS periodo,
                   Grupo_DRE, Subgrupo_DRE, SUM(Valor) AS valor
            FROM DRE
            WHERE {' AND '.join(dre_c)}
            GROUP BY periodo, Grupo_DRE, Subgrupo_DRE
            ORDER BY periodo, Grupo_DRE, Subgrupo_DRE
        """, dre_p).fetchall()

        # --- Build structures ---
        receita_by_p = {r['periodo']: r['valor'] or 0 for r in receita_rows if r['periodo']}

        # tree: {grupo: {subgrupo: {periodo: valor}}}
        tree = {}
        for r in dre_rows:
            p = r['periodo']
            g = r['Grupo_DRE']    or '(sem grupo)'
            s = r['Subgrupo_DRE'] or '(sem subgrupo)'
            v = r['valor']        or 0
            tree.setdefault(g, {}).setdefault(s, {})[p] = \
                tree.get(g, {}).get(s, {}).get(p, 0) + v

        desp_by_p = {}
        for r in dre_rows:
            if r['periodo']:
                desp_by_p[r['periodo']] = desp_by_p.get(r['periodo'], 0) + (r['valor'] or 0)

        periodos = sorted(set(list(receita_by_p.keys()) +
                              [r['periodo'] for r in dre_rows if r['periodo']]))

        total_receita  = sum(receita_by_p.values())
        total_despesas = sum(desp_by_p.values())

        chart = [{
            'periodo':   p,
            'receita':   receita_by_p.get(p, 0),
            'despesas':  desp_by_p.get(p, 0),
            'resultado': receita_by_p.get(p, 0) - desp_by_p.get(p, 0),
        } for p in periodos]

        return jsonify({
            'periodos':           periodos,
            'kpis': {
                'receita':   total_receita,
                'despesas':  total_despesas,
                'resultado': total_receita - total_despesas,
            },
            'chart':              chart,
            'receita_by_periodo': receita_by_p,
            'tree':               tree,
        })
    except sqlite3.Error as e:
        logger.error("Erro DRE report: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /summary
# ---------------------------------------------------------------------------

@dre_bp.route('/summary')
@login_required
def api_dre_summary():
    conn = get_db()
    try:
        ano      = request.args.get('ano', '')
        grupo    = request.args.get('grupo', '')
        subgrupo = request.args.get('subgrupo', '')
        situacao = request.args.get('situacao', '')

        conditions, params = _build_where(ano, grupo, subgrupo, situacao)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        kpis = conn.execute(f"""
            SELECT
                SUM(Valor)                                                    AS total,
                SUM(CASE WHEN Situacao = 'Confirmado' THEN Valor ELSE 0 END) AS confirmado,
                SUM(CASE WHEN Situacao = 'Em aberto'  THEN Valor ELSE 0 END) AS em_aberto,
                COUNT(*)                                                      AS lancamentos
            FROM DRE {where}
        """, params).fetchone()

        rows = conn.execute(f"""
            SELECT Grupo_DRE, Subgrupo_DRE, Mes, SUM(Valor) AS Total
            FROM DRE {where}
            GROUP BY Grupo_DRE, Subgrupo_DRE, Mes
            ORDER BY Grupo_DRE, Subgrupo_DRE, Mes
        """, params).fetchall()

        # Build hierarchical tree: {grupo: {subgrupo: {mes: valor}}}
        tree = {}
        for r in rows:
            g = r['Grupo_DRE']    or '(sem grupo)'
            s = r['Subgrupo_DRE'] or '(sem subgrupo)'
            m = r['Mes']
            v = r['Total'] or 0
            tree.setdefault(g, {}).setdefault(s, {})[m] = v

        return jsonify({
            "kpis": {
                "total":        kpis['total']        or 0,
                "confirmado":   kpis['confirmado']   or 0,
                "em_aberto":    kpis['em_aberto']    or 0,
                "lancamentos":  kpis['lancamentos']  or 0,
            },
            "tree": tree,
        })
    except sqlite3.Error as e:
        logger.error("Erro DRE summary: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /data  (paginado)
# ---------------------------------------------------------------------------

@dre_bp.route('/data')
@login_required
def api_dre_data():
    conn = get_db()
    try:
        page       = max(1, int(request.args.get('page', 1)))
        per_page   = min(200, max(1, int(request.args.get('per_page', 50))))
        start_date = request.args.get('start_date', '')
        end_date   = request.args.get('end_date', '')
        grupo      = request.args.get('grupo', '')
        subgrupo   = request.args.get('subgrupo', '')
        situacao   = request.args.get('situacao', '')
        search     = request.args.get('search', '').strip()

        conditions, params = [], []
        if start_date:
            conditions.append("Data_Competencia >= ?"); params.append(start_date)
        if end_date:
            conditions.append("Data_Competencia <= ?"); params.append(end_date)
        if grupo:
            conditions.append("Grupo_DRE = ?"); params.append(grupo)
        if subgrupo:
            conditions.append("Subgrupo_DRE = ?"); params.append(subgrupo)
        if situacao:
            conditions.append("Situacao = ?"); params.append(situacao)
        if search:
            conditions.append(
                "(Fornecedor LIKE ? OR Plano_de_Contas LIKE ? OR Centro_de_Custo LIKE ?)"
            )
            params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

        where  = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        offset = (page - 1) * per_page

        total = conn.execute(f"SELECT COUNT(*) FROM DRE {where}", params).fetchone()[0]
        rows  = conn.execute(
            f"SELECT * FROM DRE {where} ORDER BY Ano DESC, Mes DESC, id DESC LIMIT ? OFFSET ?",
            params + [per_page, offset],
        ).fetchall()

        return jsonify({
            "data":     [dict(r) for r in rows],
            "total":    total,
            "page":     page,
            "per_page": per_page,
            "pages":    math.ceil(total / per_page) if total else 0,
        })
    except sqlite3.Error as e:
        logger.error("Erro DRE data: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# POST /upload
# ---------------------------------------------------------------------------

@dre_bp.route('/upload', methods=['POST'])
@login_required
def api_dre_upload():
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "Nome de arquivo vazio"}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('xlsx', 'csv', 'pdf'):
        return jsonify({"error": "Formato não suportado. Use xlsx, csv ou pdf."}), 400

    try:
        rows = _parse_file(f, ext)
    except Exception as e:
        logger.error("Erro ao parsear arquivo DRE (%s): %s", ext, e, exc_info=True)
        return jsonify({"error": f"Erro ao ler arquivo: {e}"}), 400

    if not rows:
        return jsonify({"error": "Nenhum dado encontrado no arquivo"}), 400

    conn = get_db()
    try:
        conn.execute("DELETE FROM DRE")
        conn.executemany("""
            INSERT INTO DRE (
                Ano, Mes, Ano_Mes, Grupo_DRE, Subgrupo_DRE, Plano_de_Contas,
                Centro_de_Custo, Fornecedor, CNPJ, Situacao,
                Data_Competencia, Data_Vencimento, Data_Confirmacao,
                Valor, NFe, Cod_Lancamento, Loja, Observacao
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, rows)
        conn.commit()
        logger.info("DRE upload: %d registros inseridos", len(rows))
        return jsonify({"success": True, "inserted": len(rows)})
    except sqlite3.Error as e:
        conn.rollback()
        logger.error("Erro ao inserir DRE: %s", e, exc_info=True)
        return jsonify({"error": f"Erro no banco: {e}"}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# GET /auxiliar  — KPIs de negócio: Churn, CAC, LTV, ARPU, MRR, Ops
# ---------------------------------------------------------------------------

@dre_bp.route('/auxiliar')
@login_required
def api_dre_auxiliar():
    import re as _re
    conn = get_db()
    try:
        start_date    = request.args.get('start_date', '')
        end_date      = request.args.get('end_date',   '')
        cac_subgrupos  = [s.strip() for s in request.args.get('cac_subgrupos', '').split('||') if s.strip()]

        # Filtros de status — ambos obrigatórios; padrão = Ativo/Ativo
        _raw_st_c = request.args.get('st_contrato', 'Ativo')
        _raw_st_a = request.args.get('st_acesso',   'Ativo')
        st_contrato = [s.strip() for s in _raw_st_c.split(',') if s.strip()] or ['Ativo']
        st_acesso   = [s.strip() for s in _raw_st_a.split(',') if s.strip()] or ['Ativo']

        # ------------------------------------------------------------------
        # 1. Clientes ativos: Status_contrato E Status_acesso selecionáveis pelo usuário
        # ------------------------------------------------------------------
        st_c_ph = ','.join(['?'] * len(st_contrato))
        st_a_ph = ','.join(['?'] * len(st_acesso))
        active_q = conn.execute(
            f"SELECT COUNT(*) FROM Contratos "
            f"WHERE Status_contrato IN ({st_c_ph}) AND Status_acesso IN ({st_a_ph})",
            st_contrato + st_acesso,
        ).fetchone()[0] or 0

        total_cadastrados = conn.execute("SELECT COUNT(*) FROM Clientes").fetchone()[0] or 0

        # Valores disponíveis para os filtros de status
        avail_st_contrato = [r[0] for r in conn.execute(
            "SELECT DISTINCT Status_contrato FROM Contratos WHERE Status_contrato IS NOT NULL ORDER BY Status_contrato"
        ).fetchall()]
        avail_st_acesso = [r[0] for r in conn.execute(
            "SELECT DISTINCT Status_acesso FROM Contratos WHERE Status_acesso IS NOT NULL ORDER BY Status_acesso"
        ).fetchall()]

        # ------------------------------------------------------------------
        # 2. MRR por mês (Contas_a_Receber)
        # ------------------------------------------------------------------
        car_c = ["Status = 'Recebido'", "Data_pagamento IS NOT NULL", "Data_pagamento != ''"]
        car_p = []
        if start_date: car_c.append("Data_pagamento >= ?"); car_p.append(start_date)
        if end_date:   car_c.append("Data_pagamento <= ?"); car_p.append(end_date)

        mrr_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', Data_pagamento) AS periodo, SUM(Valor_recebido) AS mrr
            FROM Contas_a_Receber
            WHERE {' AND '.join(car_c)}
            GROUP BY periodo ORDER BY periodo
        """, car_p).fetchall()
        mrr_by_p = {r['periodo']: (r['mrr'] or 0) for r in mrr_rows if r['periodo']}

        # A receber por período (data de vencimento)
        ar_c = ["Status = 'A receber'", "Vencimento IS NOT NULL", "Vencimento != ''"]
        ar_p = []
        if start_date: ar_c.append("Vencimento >= ?"); ar_p.append(start_date)
        if end_date:   ar_c.append("Vencimento <= ?"); ar_p.append(end_date)
        ar_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', Vencimento) AS periodo, SUM(Valor_aberto) AS total
            FROM Contas_a_Receber
            WHERE {' AND '.join(ar_c)}
            GROUP BY periodo ORDER BY periodo
        """, ar_p).fetchall()
        ar_by_p = {r['periodo']: (r['total'] or 0) for r in ar_rows if r['periodo']}

        # ------------------------------------------------------------------
        # 3. Churn por mês (Contratos cancelados/inativos)
        # ------------------------------------------------------------------
        churn_c = [
            "(Status_contrato = 'Inativo' OR Status_contrato = 'Negativado' OR Status_contrato = 'Cancelado' OR Status_contrato = 'Desistente')",
            "Data_cancelamento IS NOT NULL", "Data_cancelamento != ''",
        ]
        churn_p = []
        if start_date: churn_c.append("Data_cancelamento >= ?"); churn_p.append(start_date)
        if end_date:   churn_c.append("Data_cancelamento <= ?"); churn_p.append(end_date)

        _churn_where = ' AND '.join(churn_c)
        churn_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', Data_cancelamento) AS periodo, COUNT(*) AS cnt
            FROM (
                SELECT Data_cancelamento FROM Contratos WHERE {_churn_where}
                UNION ALL
                SELECT Data_cancelamento FROM Contratos_Negativacao WHERE {_churn_where}
            ) t
            GROUP BY periodo ORDER BY periodo
        """, churn_p + churn_p).fetchall()
        churn_by_p  = {r['periodo']: r['cnt'] for r in churn_rows if r['periodo']}
        total_churn = sum(churn_by_p.values())

        # ------------------------------------------------------------------
        # 3b. Negativações por mês (usando Data_negativa_o)
        # Contratos: filtra Status=Negativado (evita datas históricas de Filial 2 já resolvidas)
        # Contratos_Negativacao: sem filtro de status — todos os contratos lá são por definição
        #   negados (Filial 4), e o status pode mudar após o pagamento sem invalidar a negativação
        # ------------------------------------------------------------------
        neg_c_c = ["Status_contrato = 'Negativado'", "Data_negativa_o IS NOT NULL", "Data_negativa_o != ''"]
        neg_c_n = ["Data_negativa_o IS NOT NULL", "Data_negativa_o != ''"]
        neg_p_c, neg_p_n = [], []
        if start_date:
            neg_c_c.append("Data_negativa_o >= ?"); neg_p_c.append(start_date)
            neg_c_n.append("Data_negativa_o >= ?"); neg_p_n.append(start_date)
        if end_date:
            neg_c_c.append("Data_negativa_o <= ?"); neg_p_c.append(end_date)
            neg_c_n.append("Data_negativa_o <= ?"); neg_p_n.append(end_date)

        neg_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', Data_negativa_o) AS periodo, COUNT(*) AS cnt
            FROM (
                SELECT Data_negativa_o FROM Contratos
                  WHERE {' AND '.join(neg_c_c)}
                UNION ALL
                SELECT Data_negativa_o FROM Contratos_Negativacao
                  WHERE {' AND '.join(neg_c_n)}
            ) t
            GROUP BY periodo ORDER BY periodo
        """, neg_p_c + neg_p_n).fetchall()
        neg_by_p  = {r['periodo']: r['cnt'] for r in neg_rows if r['periodo']}
        total_neg = sum(neg_by_p.values())

        # Churn por motivo
        _MOTIVO_MAP = {
            1:  'Alteração de contrato',
            2:  'Cancel. renegociação',
            3:  'A pedido do cliente',
            4:  'Pendência financeira',
            5:  'Migração de plano',
            8:  'Migração de vencimento',
            9:  'Cancelamento',
            10: 'Insatisfação',
            11: 'Mudança de endereço',
            12: 'Dificuldades financeiras',
            13: 'Viagem',
            14: 'Término de contrato',
            15: 'Suspensão temporária',
        }
        motivo_rows = conn.execute(f"""
            SELECT motivo, COUNT(*) AS cnt
            FROM (
                SELECT Motivo_cancelamento AS motivo FROM Contratos WHERE {_churn_where}
                UNION ALL
                SELECT Motivo_cancelamento AS motivo FROM Contratos_Negativacao WHERE {_churn_where}
            ) t
            GROUP BY motivo ORDER BY cnt DESC
        """, churn_p + churn_p).fetchall()
        def _motivo_label(raw):
            """Motivo_cancelamento é armazenado como TEXT, mas o mapa usa chaves int."""
            if raw is None:
                return None, 'Sem informação'
            try:
                mid = int(raw)
            except (ValueError, TypeError):
                mid = None
            label = _MOTIVO_MAP.get(mid, f'Código {raw}')
            return mid, label

        churn_by_motivo = []
        for r in motivo_rows:
            mid, label = _motivo_label(r['motivo'])
            churn_by_motivo.append({
                'motivo_id': mid,
                'motivo':    label,
                'count':     r['cnt'],
                'pct':       round(r['cnt'] / total_churn * 100, 1)
                             if total_churn > 0 else 0.0,
            })

        # ------------------------------------------------------------------
        # 4. Novas ativações por mês
        # ------------------------------------------------------------------
        new_c = ["Data_ativa_o IS NOT NULL", "Data_ativa_o != ''"]
        new_p = []
        if start_date: new_c.append("Data_ativa_o >= ?"); new_p.append(start_date)
        if end_date:   new_c.append("Data_ativa_o <= ?"); new_p.append(end_date)

        new_rows = conn.execute(f"""
            SELECT STRFTIME('%Y-%m', Data_ativa_o) AS periodo, COUNT(*) AS cnt
            FROM Contratos
            WHERE {' AND '.join(new_c)}
            GROUP BY periodo ORDER BY periodo
        """, new_p).fetchall()
        new_by_p  = {r['periodo']: r['cnt'] for r in new_rows if r['periodo']}
        total_new = sum(new_by_p.values())

        # ------------------------------------------------------------------
        # 5. Métricas derivadas do período
        # ------------------------------------------------------------------
        all_periods = sorted(
            set(list(mrr_by_p.keys()) + list(churn_by_p.keys()) + list(new_by_p.keys())
                + list(neg_by_p.keys()) + list(ar_by_p.keys()))
        )
        period_months = max(1, len(all_periods))

        last_p           = all_periods[-1] if all_periods else None
        db_recebido      = mrr_by_p.get(last_p, 0) if last_p else 0
        db_a_receber     = ar_by_p.get(last_p, 0) if last_p else 0
        db_receita_bruta = db_recebido + db_a_receber
        mrr_current      = db_receita_bruta
        arpu             = mrr_current / active_q if active_q > 0 else 0

        # Totais do período inteiro (BD)
        period_db_recebido      = sum(mrr_by_p.values())
        period_db_a_receber     = sum(ar_by_p.values())
        period_db_receita_bruta = period_db_recebido + period_db_a_receber

        avg_monthly_churn    = total_churn / period_months
        churn_rate_pct       = (avg_monthly_churn / active_q * 100) if active_q > 0 else 0
        cancel_rate_pct      = (total_churn / active_q * 100) if active_q > 0 else 0
        avg_monthly_neg      = total_neg / period_months
        neg_rate_pct         = (total_neg / active_q * 100) if active_q > 0 else 0
        total_combined       = total_churn + total_neg
        avg_monthly_combined = total_combined / period_months
        combined_rate_pct    = (total_combined / active_q * 100) if active_q > 0 else 0

        # Taxa LTV (últimos 12 meses, janela fixa) — evita distorção com períodos longos.
        # Negativado = saída permanente nesta operação (novo contrato se quiser voltar).
        # Metodologia: cancel/inativo/desistente + negativados de Contratos
        #              + negativações de Contratos_Negativacao (Filial 4).
        _ltv_combined_row = conn.execute("""
            SELECT COUNT(*) AS cnt FROM (
                SELECT 1 FROM Contratos
                  WHERE Status_contrato IN ('Cancelado','Inativo','Desistente')
                    AND Data_cancelamento IS NOT NULL AND Data_cancelamento != ''
                    AND Data_cancelamento >= DATE('now', '-12 months')
                UNION ALL
                SELECT 1 FROM Contratos
                  WHERE Status_contrato = 'Negativado'
                    AND Data_negativa_o IS NOT NULL AND Data_negativa_o != ''
                    AND Data_negativa_o >= DATE('now', '-12 months')
                UNION ALL
                SELECT 1 FROM Contratos_Negativacao
                  WHERE Data_negativa_o IS NOT NULL AND Data_negativa_o != ''
                    AND Data_negativa_o >= DATE('now', '-12 months')
            ) t
        """).fetchone()
        _ltv_combined_12 = (_ltv_combined_row['cnt'] or 0)
        _ltv_rate_pct    = (_ltv_combined_12 / 12.0 / active_q * 100) if active_q > 0 else combined_rate_pct

        # ------------------------------------------------------------------
        # 6. CAC — custo dos subgrupos selecionados ÷ novas ativações
        # ------------------------------------------------------------------
        cac_cost = 0.0
        if cac_subgrupos:
            ph     = ','.join(['?'] * len(cac_subgrupos))
            cac_c2 = [f"Subgrupo_DRE IN ({ph})"]
            cac_p2 = list(cac_subgrupos)
            if start_date: cac_c2.append("Data_Competencia >= ?"); cac_p2.append(start_date)
            if end_date:   cac_c2.append("Data_Competencia <= ?"); cac_p2.append(end_date)
            row = conn.execute(
                f"SELECT SUM(Valor) FROM DRE WHERE {' AND '.join(cac_c2)}", cac_p2
            ).fetchone()
            cac_cost = row[0] or 0.0

        cac      = cac_cost / total_new if total_new > 0 else 0.0
        avg_life = 1.0 / (_ltv_rate_pct / 100.0) if _ltv_rate_pct > 0 else 0.0
        ltv      = arpu * avg_life
        ltv_cac  = ltv / cac if cac > 0 else 0.0
        payback  = cac / arpu if arpu > 0 else 0.0

        # ------------------------------------------------------------------
        # 6b. LTV Real — permanência e receita real por contrato
        #     Conta apenas os meses em que o contrato efetivamente pagou
        #     (Contas_a_Receber, Status='Recebido'). Janela: histórico completo,
        #     sem filtro de data (queremos o total real de permanência de cada cliente).
        # ------------------------------------------------------------------
        _ltv_real_row = conn.execute("""
            SELECT AVG(months_count)   AS avg_months,
                   AVG(contract_total) AS avg_revenue
            FROM (
                SELECT ID_contrato_principal,
                       COUNT(DISTINCT STRFTIME('%Y-%m', Data_pagamento)) AS months_count,
                       SUM(Valor_recebido)                               AS contract_total
                FROM Contas_a_Receber
                WHERE Status = 'Recebido'
                  AND Data_pagamento IS NOT NULL AND Data_pagamento != ''
                  AND ID_contrato_principal IS NOT NULL AND ID_contrato_principal > 0
                GROUP BY ID_contrato_principal
            )
        """).fetchone()
        ltv_real_months  = float(_ltv_real_row['avg_months']  or 0) if _ltv_real_row else 0.0
        ltv_real         = float(_ltv_real_row['avg_revenue'] or 0) if _ltv_real_row else 0.0

        # ------------------------------------------------------------------
        # 7. PMR (Prazo Médio de Recebimento)
        # ------------------------------------------------------------------
        pmr_c = [
            "Data_pagamento IS NOT NULL", "Data_pagamento != ''",
            "Vencimento IS NOT NULL", "Vencimento != ''", "Status = 'Recebido'",
        ]
        pmr_p = []
        if start_date: pmr_c.append("Data_pagamento >= ?"); pmr_p.append(start_date)
        if end_date:   pmr_c.append("Data_pagamento <= ?"); pmr_p.append(end_date)

        pmr_row = conn.execute(f"""
            SELECT
                AVG(JULIANDAY(Data_pagamento) - JULIANDAY(Vencimento))                       AS pmr_all,
                AVG(CASE WHEN JULIANDAY(Data_pagamento) > JULIANDAY(Vencimento)
                    THEN JULIANDAY(Data_pagamento) - JULIANDAY(Vencimento)
                    ELSE NULL END)                                                             AS pmr_late
            FROM Contas_a_Receber
            WHERE {' AND '.join(pmr_c)}
        """, pmr_p).fetchone()
        pmr_days      = round(pmr_row['pmr_all']  or 0, 1)
        pmr_late_days = round(pmr_row['pmr_late'] or 0, 1)

        # ------------------------------------------------------------------
        # 8. Inadimplência
        # ------------------------------------------------------------------
        inadimplencia_valor = 0.0
        inadimplencia_pct   = 0.0
        try:
            inad_start = start_date or '2020-01-01'
            inad_row = conn.execute("""
                SELECT COUNT(*) AS cnt, SUM(Valor_aberto) AS valor
                FROM Contas_a_Receber
                WHERE Status = 'A receber' AND Vencimento < DATE('now') AND Vencimento >= ?
            """, (inad_start,)).fetchone()
            inad_total = conn.execute("""
                SELECT COUNT(*) FROM Contas_a_Receber
                WHERE Status = 'A receber' AND Vencimento >= ?
            """, (inad_start,)).fetchone()[0] or 1
            inadimplencia_valor = inad_row['valor'] or 0.0
            inadimplencia_pct   = ((inad_row['cnt'] or 0) / inad_total * 100)
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 9. OS: truck rolls, MTTR, instalação, uptime
        # ------------------------------------------------------------------
        os_c = ["Abertura IS NOT NULL"]
        os_p = []
        if start_date: os_c.append("Abertura >= ?"); os_p.append(start_date)
        if end_date:   os_c.append("Abertura <= ?"); os_p.append(end_date)
        where_os = ' AND '.join(os_c)

        _truck_types = (
            "UPPER(Assunto) LIKE '%MUDAN%PONTO%'"
            " OR UPPER(Assunto) LIKE '%MANUTEN%FIBRA%'"
            " OR UPPER(Assunto) LIKE '%VISITA%TECNI%'"
            " OR UPPER(Assunto) LIKE '%VISITA%TÉCNI%'"
            " OR UPPER(Assunto) LIKE '%RETORNO%MANUTEN%'"
            " OR UPPER(Assunto) LIKE '%REATIVA%CLIENTE%'"
        )
        os_total = conn.execute(
            f"SELECT COUNT(*) FROM OS WHERE ({_truck_types}) AND {where_os}", os_p
        ).fetchone()[0] or 0
        truck_per_client = os_total / active_q if active_q > 0 else 0.0

        manut_row = conn.execute(f"""
            SELECT COUNT(*) AS cnt,
                   AVG(CASE WHEN Fechamento IS NOT NULL AND Abertura IS NOT NULL
                       THEN (JULIANDAY(Fechamento) - JULIANDAY(Abertura)) * 24
                       ELSE NULL END) AS avg_h
            FROM OS
            WHERE (UPPER(Assunto) LIKE '%MANUTEN%'
                OR UPPER(Assunto) LIKE '%OSCILA%'
                OR UPPER(Assunto) LIKE '%SEM CONEX%'
                OR UPPER(Assunto) LIKE '%SEM INTERNET%'
                OR UPPER(Assunto) LIKE '%FALHA%')
              AND {where_os}
        """, os_p).fetchone()
        mttr_hours  = round(manut_row['avg_h'] or 0, 1)
        manut_count = manut_row['cnt'] or 0

        total_period_h = period_months * 30 * 24
        lost_h         = mttr_hours * manut_count
        uptime_pct     = round(max(0.0, (1 - lost_h / max(1, total_period_h)) * 100), 2)

        inst_row = conn.execute(f"""
            SELECT AVG(CASE WHEN Fechamento IS NOT NULL AND Abertura IS NOT NULL
                THEN (JULIANDAY(Fechamento) - JULIANDAY(Abertura)) * 24
                ELSE NULL END) AS avg_h
            FROM OS
            WHERE UPPER(Assunto) LIKE '%INSTALA%' AND {where_os}
        """, os_p).fetchone()
        tempo_instalacao = round(inst_row['avg_h'] or 0, 1)

        # ------------------------------------------------------------------
        # 10. Resolução 1º contato (Atendimentos)
        # ------------------------------------------------------------------
        resolucao_1c   = 0.0
        total_aten     = 0
        com_dados      = 0
        try:
            aten_c, aten_p = [], []
            if start_date: aten_c.append("Criado_em >= ?"); aten_p.append(start_date)
            if end_date:   aten_c.append("Criado_em <= ?"); aten_p.append(end_date)
            where_at = ("WHERE " + " AND ".join(aten_c)) if aten_c else ""
            aten_row = conn.execute(f"""
                SELECT COUNT(*) AS total,
                       SUM(CASE WHEN Msgs_cliente IS NOT NULL AND Msgs_cliente <= 1 THEN 1 ELSE 0 END) AS first,
                       SUM(CASE WHEN Msgs_cliente IS NOT NULL THEN 1 ELSE 0 END) AS com_dados
                FROM Atendimentos {where_at}
            """, aten_p).fetchone()
            total_aten = aten_row['total'] or 0
            com_dados  = aten_row['com_dados'] or 0
            resolucao_1c = round(
                ((aten_row['first'] or 0) / com_dados * 100) if com_dados > 0 else 0.0, 1
            )
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 11. OPEX por cliente / mês
        # ------------------------------------------------------------------
        opex_c = ["Data_Competencia IS NOT NULL", "Data_Competencia != ''"]
        opex_p = []
        if start_date: opex_c.append("Data_Competencia >= ?"); opex_p.append(start_date)
        if end_date:   opex_c.append("Data_Competencia <= ?"); opex_p.append(end_date)
        total_opex    = conn.execute(
            f"SELECT SUM(Valor) FROM DRE WHERE {' AND '.join(opex_c)}", opex_p
        ).fetchone()[0] or 0.0
        opex_per_client = total_opex / (active_q * period_months) if (active_q * period_months) > 0 else 0.0

        # ------------------------------------------------------------------
        # 12. Custo por Mbps (extrai velocidade do campo Descri_o)
        # ------------------------------------------------------------------
        mbps_rows = conn.execute("""
            SELECT Descri_o, COUNT(*) AS cnt
            FROM Contratos
            WHERE Status_contrato = 'Ativo' AND Descri_o IS NOT NULL AND Descri_o != ''
            GROUP BY Descri_o
        """).fetchall()
        total_mbps = 0
        for row in mbps_rows:
            m = _re.search(r'(\d+)\s*M', str(row['Descri_o']), _re.IGNORECASE)
            if m:
                total_mbps += int(m.group(1)) * (row['cnt'] or 0)
        custo_por_mbps = round(total_opex / total_mbps, 4) if total_mbps > 0 else 0.0

        # ------------------------------------------------------------------
        # 13. Taxa de penetração por cidade
        # ------------------------------------------------------------------
        taxa_penetracao = round(active_q / total_cadastrados * 100, 1) if total_cadastrados > 0 else 0.0
        city_pen = []
        try:
            city_rows = conn.execute("""
                SELECT CO.Cidade AS cidade, COUNT(*) AS ativos,
                       (SELECT COUNT(*) FROM Clientes CL WHERE CL.Cidade = CO.Cidade) AS cadastrados
                FROM Contratos CO
                WHERE CO.Status_contrato = 'Ativo' AND CO.Cidade IS NOT NULL AND CO.Cidade != ''
                  AND CO.Cidade NOT GLOB '[0-9]*'
                GROUP BY CO.Cidade
                HAVING cadastrados > 0
                ORDER BY ativos DESC
                LIMIT 25
            """).fetchall()
            city_pen = [
                {
                    'cidade':      r['cidade'],
                    'ativos':      r['ativos'],
                    'cadastrados': r['cadastrados'],
                    'pct':         round(r['ativos'] / r['cadastrados'] * 100, 1)
                                   if r['cadastrados'] > 0 else 0.0,
                }
                for r in city_rows
            ]
        except Exception:
            pass

        # ------------------------------------------------------------------
        # 14. Subgrupos disponíveis para CAC e status selecionados
        # ------------------------------------------------------------------
        avail_subgrupos = [
            r[0] for r in conn.execute("""
                SELECT DISTINCT sub FROM (
                    SELECT Subgrupo_DRE AS sub FROM DRE WHERE Subgrupo_DRE IS NOT NULL
                    UNION
                    SELECT SubgrupoDRE  AS sub FROM GC_Lancamentos WHERE SubgrupoDRE IS NOT NULL
                ) ORDER BY sub
            """).fetchall()
        ]

        # ------------------------------------------------------------------
        # 15. Dados financeiros do Excel (GC_*)
        # ------------------------------------------------------------------
        gestao = {}
        try:
            gc_conds, gc_params = [], []
            if start_date: gc_conds.append("AnoMes >= ?"); gc_params.append(start_date[:7])
            if end_date:   gc_conds.append("AnoMes <= ?"); gc_params.append(end_date[:7])
            wg = (" AND " + " AND ".join(gc_conds)) if gc_conds else ""

            dre_last = conn.execute(f"""
                SELECT ReceitaBruta, Recebido, AReceber, TotalDespesas, Resultado, Margem
                FROM GC_DRE_Completo WHERE 1=1{wg}
                ORDER BY AnoMes DESC LIMIT 1
            """, gc_params).fetchone()
            dre_agg = conn.execute(f"""
                SELECT SUM(ReceitaBruta) AS rb_total, SUM(Recebido) AS rec_total,
                       SUM(AReceber) AS ar_total, SUM(TotalDespesas) AS desp_total,
                       SUM(Resultado) AS res_total, COUNT(*) AS n
                FROM GC_DRE_Completo WHERE 1=1{wg}
            """, gc_params).fetchone()

            if dre_last:
                gestao.update({
                    'gc_receita_bruta': float(dre_last['ReceitaBruta'] or 0),
                    'gc_recebido':      float(dre_last['Recebido']     or 0),
                    'gc_a_receber':     float(dre_last['AReceber']     or 0),
                    'gc_total_desp':    float(dre_last['TotalDespesas']or 0),
                    'gc_resultado':     float(dre_last['Resultado']    or 0),
                    'gc_margem':        float(dre_last['Margem']       or 0),
                })
            if dre_agg and (dre_agg['n'] or 0) > 0:
                gestao.update({
                    'gc_rb_total':   float(dre_agg['rb_total']   or 0),
                    'gc_rec_total':  float(dre_agg['rec_total']  or 0),
                    'gc_ar_total':   float(dre_agg['ar_total']   or 0),
                    'gc_desp_total': float(dre_agg['desp_total'] or 0),
                    'gc_res_total':  float(dre_agg['res_total']  or 0),
                    'gc_n_meses':    int(dre_agg['n']),
                })

            # Instalações: OS de ativação finalizadas no período (banco de dados)
            os_inst_c = [
                "Assunto = 'INSTALAÇÃO DE FIBRA'",
                "Status = 'Finalizada'",
                "Fechamento IS NOT NULL", "Fechamento != ''",
            ]
            os_inst_p = []
            if start_date: os_inst_c.append("Fechamento >= ?"); os_inst_p.append(start_date)
            if end_date:   os_inst_c.append("Fechamento <= ?"); os_inst_p.append(end_date + ' 23:59:59')
            os_inst_row = conn.execute(
                f"SELECT COUNT(*) AS cnt FROM OS WHERE {' AND '.join(os_inst_c)}", os_inst_p
            ).fetchone()
            gc_n = int(os_inst_row['cnt'] or 0) if os_inst_row else 0

            # Custo CAC: sempre via GC_Lancamentos filtrado pelos subgrupos do configurador
            # Sem subgrupos → CAC não pode ser calculado corretamente
            gestao.update({
                'gc_instalacoes': gc_n,
                'gc_cac_source':  'none',
                'gc_cac_total':   0.0,
                'gc_cac_unit':    0.0,
            })
            if cac_subgrupos:
                ph   = ','.join(['?'] * len(cac_subgrupos))
                lc_c = [f"SubgrupoDRE IN ({ph})"]
                lc_p = list(cac_subgrupos)
                if start_date: lc_c.append("AnoMes >= ?"); lc_p.append(start_date[:7])
                if end_date:   lc_c.append("AnoMes <= ?"); lc_p.append(end_date[:7])
                lc_row = conn.execute(
                    f"SELECT SUM(Valor) AS total FROM GC_Lancamentos WHERE {' AND '.join(lc_c)}", lc_p
                ).fetchone()
                gc_t = float(lc_row['total'] or 0) if lc_row else 0.0
                gestao.update({
                    'gc_cac_total':  gc_t,
                    'gc_cac_unit':   gc_t / gc_n if gc_n > 0 else 0.0,
                    'gc_cac_source': 'subgrupos',
                })

            dfc_last = conn.execute(f"""
                SELECT SaldoAcumulado FROM GC_DFC_Mensal WHERE 1=1{wg}
                ORDER BY AnoMes DESC LIMIT 1
            """, gc_params).fetchone()
            dfc_agg = conn.execute(f"""
                SELECT AVG(Entradas) AS ent, AVG(TotalSaidas) AS sai,
                       AVG(SaldoPeriodo) AS sp, SUM(TotalSaidas) AS sai_total, COUNT(*) AS n
                FROM GC_DFC_Mensal WHERE 1=1{wg}
            """, gc_params).fetchone()
            if dfc_agg and (dfc_agg['n'] or 0) > 0:
                gestao.update({
                    'gc_entradas':      float(dfc_agg['ent']       or 0),
                    'gc_saidas':        float(dfc_agg['sai']       or 0),
                    'gc_saldo_periodo': float(dfc_agg['sp']        or 0),
                    'gc_saldo_acum':    float(dfc_last['SaldoAcumulado'] or 0) if dfc_last else 0,
                    'gc_saidas_total':  float(dfc_agg['sai_total'] or 0),
                })

        except Exception:
            pass

        # Receita Bruta sempre do BD (Contas_a_Receber); Excel usado apenas para Despesas/DFC
        if gestao.get('gc_cac_unit', 0) > 0:
            cac      = gestao['gc_cac_unit']
            cac_cost = gestao.get('gc_cac_total', cac_cost)
            ltv_cac  = ltv / cac if cac > 0 else 0.0
            payback  = cac / arpu if arpu > 0 else 0.0

        if gestao.get('gc_saidas_total', 0) > 0:
            total_opex      = gestao['gc_saidas_total']
            opex_per_client = total_opex / (active_q * period_months) if (active_q * period_months) > 0 else 0.0
            custo_por_mbps  = round(total_opex / total_mbps, 4) if total_mbps > 0 else 0.0

        # Resultado e margem acumulados do período (BD recebido × Excel despesas)
        period_gc_desp_total    = gestao.get('gc_desp_total', 0)
        period_resultado        = period_db_recebido - period_gc_desp_total
        period_margem           = (period_resultado / period_db_recebido * 100) if period_db_recebido > 0 else 0.0

        # ------------------------------------------------------------------
        # 16. Dados de tendência para gráfico
        # ------------------------------------------------------------------
        trend = [
            {
                'periodo':    p,
                'mrr':        mrr_by_p.get(p, 0),
                'churn':      churn_by_p.get(p, 0),
                'neg':        neg_by_p.get(p, 0),
                'novos':      new_by_p.get(p, 0),
                'churn_rate': round(churn_by_p.get(p, 0) / active_q * 100, 2)
                              if active_q > 0 else 0.0,
            }
            for p in all_periods
        ]

        return jsonify({
            'kpis': {
                'mrr':                round(mrr_current, 2),
                'arpu':               round(arpu, 2),
                'active_clients':     int(active_q),
                'total_cadastrados':  int(total_cadastrados),
                'churn_rate':         round(churn_rate_pct, 2),
                'cancel_rate':        round(cancel_rate_pct, 2),
                'churn_count':        int(round(avg_monthly_churn)),
                'churn_total':        int(total_churn),
                'neg_count':          int(round(avg_monthly_neg)),
                'neg_total':          int(total_neg),
                'neg_rate':           round(neg_rate_pct, 2),
                'combined_count':     int(round(avg_monthly_combined)),
                'combined_total':     int(total_combined),
                'combined_rate':      round(combined_rate_pct, 2),
                'ltv_churn_rate':     round(_ltv_rate_pct, 2),
                'db_receita_bruta':         round(db_receita_bruta, 2),
                'db_recebido':              round(db_recebido, 2),
                'db_a_receber':             round(db_a_receber, 2),
                'period_db_recebido':       round(period_db_recebido, 2),
                'period_db_a_receber':      round(period_db_a_receber, 2),
                'period_db_receita_bruta':  round(period_db_receita_bruta, 2),
                'period_gc_desp_total':     round(period_gc_desp_total, 2),
                'period_resultado':         round(period_resultado, 2),
                'period_margem':            round(period_margem, 2),
                'new_clients':        int(total_new),
                'cac':                round(cac, 2),
                'cac_cost':           round(cac_cost, 2),
                'ltv':                round(ltv, 2),
                'ltv_real':           round(ltv_real, 2),
                'ltv_real_months':    round(ltv_real_months, 1),
                'ltv_cac':            round(ltv_cac, 2),
                'payback_months':     round(payback, 2),
                'pmr_days':           pmr_days,
                'pmr_late_days':      pmr_late_days,
                'inadimplencia_pct':  round(inadimplencia_pct, 1),
                'inadimplencia_valor':round(inadimplencia_valor, 2),
                'truck_rolls':        int(os_total),
                'truck_per_client':   round(truck_per_client, 2),
                'mttr_hours':         mttr_hours,
                'uptime_pct':         uptime_pct,
                'tempo_instalacao':   tempo_instalacao,
                'resolucao_1c':       resolucao_1c,
                'total_atendimentos': int(total_aten),
                'resolucao_com_dados': int(com_dados),
                'opex_per_client':    round(opex_per_client, 2),
                'total_opex':         round(total_opex, 2),
                'custo_por_mbps':     custo_por_mbps,
                'total_mbps':         int(total_mbps),
                'taxa_penetracao':    taxa_penetracao,
                # Gestão Excel
                **{gk: round(gv, 2) if isinstance(gv, float) else gv
                   for gk, gv in gestao.items()},
                'gc_has_data':        bool(gestao.get('gc_receita_bruta', 0)),
            },
            'trend':             trend,
            'city_penetracao':   city_pen,
            'avail_subgrupos':   avail_subgrupos,
            'period_months':     period_months,
            'churn_by_motivo':   churn_by_motivo,
            'avail_st_contrato': avail_st_contrato,
            'avail_st_acesso':   avail_st_acesso,
            'st_contrato_sel':   st_contrato,
            'st_acesso_sel':     st_acesso,
        })

    except sqlite3.Error as e:
        logger.error("Erro DRE auxiliar (SQL): %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.error("Erro DRE auxiliar: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _build_where(ano='', grupo='', subgrupo='', situacao=''):
    conditions, params = [], []
    if ano:
        try:
            conditions.append("Ano = ?")
            params.append(int(ano))
        except ValueError:
            pass
    if grupo:
        conditions.append("Grupo_DRE = ?")
        params.append(grupo)
    if subgrupo:
        conditions.append("Subgrupo_DRE = ?")
        params.append(subgrupo)
    if situacao:
        conditions.append("Situacao = ?")
        params.append(situacao)
    return conditions, params


def _parse_file(f, ext):
    if ext == 'xlsx':
        return _parse_xlsx(f)
    if ext == 'csv':
        return _parse_csv(f)
    if ext == 'pdf':
        return _parse_pdf(f)
    return []


def _parse_xlsx(f):
    import openpyxl
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active

    # Find the header row (first cell = 'Ano', case-insensitive)
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if row and str(row[0] or '').strip().lower() == 'ano':
            header_row = i
            break
    if header_row is None:
        raise ValueError("Cabeçalho 'Ano' não encontrado nas primeiras 15 linhas")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or row[0] is None:
            continue
        v = list(row[:18]) + [None] * max(0, 18 - len(row))
        rows.append(_make_tuple(v))
    return rows


def _parse_csv(f):
    import pandas as pd
    import chardet
    content = f.read()
    enc = chardet.detect(content).get('encoding') or 'utf-8'
    df  = pd.read_csv(io.BytesIO(content), encoding=enc, sep=None, engine='python')
    _normalize_df_cols(df)
    return [_make_tuple_from_series(r) for _, r in df.iterrows()]


def _parse_pdf(f):
    import pdfplumber
    rows    = []
    content = f.read()
    found_header = False
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or all(c is None for c in row):
                        continue
                    if not found_header:
                        if str(row[0] or '').strip().lower() == 'ano':
                            found_header = True
                        continue
                    if row[0]:
                        v = [str(c).strip() if c else None for c in row]
                        v += [None] * max(0, 18 - len(v))
                        rows.append(_make_tuple(v))
    if not found_header:
        raise ValueError("Cabeçalho 'Ano' não encontrado no PDF")
    return rows


# Expected column order (xlsx positional / csv after normalize)
_CSV_COL_MAP = {
    'ano': 'Ano', 'mês': 'Mes', 'mes': 'Mes',
    'ano-mês': 'Ano_Mes', 'ano-mes': 'Ano_Mes',
    'grupo dre': 'Grupo_DRE', 'subgrupo dre': 'Subgrupo_DRE',
    'plano de contas': 'Plano_de_Contas',
    'centro de custo': 'Centro_de_Custo',
    'fornecedor': 'Fornecedor', 'cnpj': 'CNPJ',
    'situação': 'Situacao', 'situacao': 'Situacao',
    'data competência': 'Data_Competencia', 'data competencia': 'Data_Competencia',
    'data vencimento': 'Data_Vencimento',
    'data confirmação': 'Data_Confirmacao', 'data confirmacao': 'Data_Confirmacao',
    'valor (r$)': 'Valor', 'valor': 'Valor',
    'nf-e': 'NFe', 'nfe': 'NFe',
    'cód. lançamento': 'Cod_Lancamento', 'cod. lancamento': 'Cod_Lancamento',
    'loja': 'Loja', 'observação': 'Observacao', 'observacao': 'Observacao',
}


def _normalize_df_cols(df):
    df.columns = [_CSV_COL_MAP.get(c.strip().lower(), c) for c in df.columns]


def _make_tuple_from_series(r):
    def _g(k): return r.get(k)
    def _d(v):
        if v is None: return None
        s = str(v)
        return s[:10] if len(s) >= 10 else None
    return (
        _safe_int(_g('Ano')),
        _safe_int(_g('Mes')),
        _str(_g('Ano_Mes')),
        _str(_g('Grupo_DRE')),
        _str(_g('Subgrupo_DRE')),
        _str(_g('Plano_de_Contas')),
        _str(_g('Centro_de_Custo')),
        _str(_g('Fornecedor')),
        _str(_g('CNPJ')),
        _str(_g('Situacao')),
        _d(_g('Data_Competencia')),
        _d(_g('Data_Vencimento')),
        _d(_g('Data_Confirmacao')),
        _safe_float(_g('Valor')),
        _str(_g('NFe')),
        _safe_int(_g('Cod_Lancamento')),
        _str(_g('Loja')),
        _str(_g('Observacao')),
    )


def _make_tuple(v):
    """Convert positional list (xlsx/pdf row) to insert tuple."""
    def _d(val):
        if val is None:
            return None
        if isinstance(val, dt):
            return val.strftime('%Y-%m-%d')
        s = str(val).strip()
        return s[:10] if len(s) >= 10 else None

    return (
        _safe_int(v[0]),   # Ano
        _safe_int(v[1]),   # Mes
        _str(v[2]),        # Ano_Mes
        _str(v[3]),        # Grupo_DRE
        _str(v[4]),        # Subgrupo_DRE
        _str(v[5]),        # Plano_de_Contas
        _str(v[6]),        # Centro_de_Custo
        _str(v[7]),        # Fornecedor
        _str(v[8]),        # CNPJ
        _str(v[9]),        # Situacao
        _d(v[10]),         # Data_Competencia
        _d(v[11]),         # Data_Vencimento
        _d(v[12]),         # Data_Confirmacao
        _safe_float(v[13]),# Valor
        _str(v[14]),       # NFe
        _safe_int(v[15]),  # Cod_Lancamento
        _str(v[16]),       # Loja
        _str(v[17]),       # Observacao
    )


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ('none', 'nan', 'nat') else None


def _safe_int(v):
    try:
        return int(float(str(v)))
    except Exception:
        return None


def _safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).strip().replace('R$', '').replace(' ', '')
        # Brazilian format: '1.234,56' → remove thousands dot, convert comma decimal
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        return float(s)
    except Exception:
        return None
