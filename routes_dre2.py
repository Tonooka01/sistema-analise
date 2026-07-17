"""
routes_dre2.py — Gestão Financeira (DRE, DFC, CAC, Lançamentos)
Importação via upload de arquivo Excel (qualquer nome).
"""

import io
import sqlite3
import traceback
from flask import Blueprint, jsonify, request
from flask_login import login_required, current_user
from database import get_db_connection as get_db
from logger import get_logger

logger = get_logger(__name__)

dre2_bp = Blueprint('dre2', __name__)


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _add_columns(conn, table, cols):
    """Adiciona colunas que não existem ainda (idempotente)."""
    for col_name, col_type in cols:
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type}")
        except Exception:
            pass


def _ensure_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS GC_DRE_Completo (
            AnoMes TEXT PRIMARY KEY,
            Ano INTEGER, Mes TEXT,
            ReceitaBruta REAL, Recebido REAL, AReceber REAL,
            CMV REAL, DespOp REAL, Encargos REAL, DespFin REAL, Outros REAL,
            TotalDespesas REAL, Resultado REAL, Margem REAL
        );
        CREATE TABLE IF NOT EXISTS GC_DFC_Mensal (
            AnoMes TEXT PRIMARY KEY,
            Ano INTEGER, Mes TEXT,
            Entradas REAL, CMV REAL, DespOp REAL, Encargos REAL, DespFin REAL, Outros REAL,
            TotalSaidas REAL, SaldoPeriodo REAL, SaldoAcumulado REAL
        );
        CREATE TABLE IF NOT EXISTS GC_CAC_Mensal (
            AnoMes TEXT PRIMARY KEY,
            Ano INTEGER, Mes TEXT,
            Comissionamento REAL, Marketing REAL, MaterialCampo REAL, ONTs REAL,
            TotalCAC REAL, NInstalacoes INTEGER, CACUnitario REAL
        );
        CREATE TABLE IF NOT EXISTS GC_Lancamentos (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            Ano INTEGER, Mes TEXT, AnoMes TEXT,
            GrupoDRE TEXT, SubgrupoDRE TEXT, PlanoContas TEXT,
            CentroCusto TEXT, Fornecedor TEXT, CNPJ TEXT,
            Situacao TEXT, DataCompetencia TEXT, DataVencimento TEXT, DataConfirmacao TEXT,
            Valor REAL, NFe TEXT, CodLancamento INTEGER, Loja TEXT, Descricao TEXT
        );
        CREATE TABLE IF NOT EXISTS GC_DRE_Estruturado (
            Ano     INTEGER,
            Campo   TEXT,
            Valor   REAL,
            PRIMARY KEY (Ano, Campo)
        );
        CREATE TABLE IF NOT EXISTS GC_CAPEX_OPEX_Sheet (
            Secao    TEXT,
            Categoria TEXT,
            Ordem    INTEGER DEFAULT 0,
            Ano      INTEGER,
            Valor    REAL,
            PRIMARY KEY (Secao, Categoria, Ano)
        );
    """)
    conn.commit()
    # Migrations — novas colunas do v9
    _add_columns(conn, 'GC_DFC_Mensal', [
        ('Pessoal', 'REAL'), ('EncargosTrabalh', 'REAL'), ('Marketing_DFC', 'REAL'),
        ('Infraestrutura', 'REAL'), ('Tecnologia', 'REAL'), ('Frota', 'REAL'),
        ('DespAdmin', 'REAL'), ('Atendimento', 'REAL'), ('Impostos', 'REAL'), ('IRPJCSLL', 'REAL'),
    ])
    _add_columns(conn, 'GC_Lancamentos', [
        ('CategoriaEstruturada', 'TEXT'), ('CapexOpex', 'TEXT'),
    ])
    conn.commit()


def _import_excel(conn, file_bytes):
    import openpyxl

    def _dt(v):
        if v is None:
            return None
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)

    def _f(v):
        if v is None:
            return 0.0
        try:
            return float(v)
        except Exception:
            return 0.0

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    conn.execute("DELETE FROM GC_DRE_Completo")
    conn.execute("DELETE FROM GC_DFC_Mensal")
    conn.execute("DELETE FROM GC_CAC_Mensal")
    conn.execute("DELETE FROM GC_Lancamentos")
    conn.execute("DELETE FROM GC_DRE_Estruturado")
    conn.execute("DELETE FROM GC_CAPEX_OPEX_Sheet")

    # --- DRE Completo ---
    ws = wb['📈 DRE Completo']
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not isinstance(row[0], int):
            continue
        conn.execute(
            "INSERT OR REPLACE INTO GC_DRE_Completo VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (row[2], row[0], row[1],
             _f(row[3]), _f(row[4]), _f(row[5]),
             _f(row[6]), _f(row[7]), _f(row[8]), _f(row[9]), _f(row[10]),
             _f(row[11]), _f(row[12]), _f(row[13]))
        )

    # --- DFC Mensal (v9: colunas ordenadas alfanumericamente pelo prefixo) ---
    # idx: 3=ENTRADAS, 4=1.CMV, 5=10.DespFin, 6=11.Outros, 7=12.Atendimento,
    #      8=2.Pessoal, 9=3.EncargosTrabalh, 10=4.Marketing, 11=5.Infra,
    #      12=6.Tecnologia, 13=7.Frota, 14=8.DespAdmin, 15=8.Impostos,
    #      16=9.IRPJ, 17=TOTAL SAÍDAS, 18=SALDO PERÍODO, 19=SALDO ACUMULADO
    ws = wb['💵 DFC Mensal']
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not isinstance(row[0], int):
            continue
        atendimento = _f(row[7])
        pessoal     = _f(row[8])
        enc_trabal  = _f(row[9])
        marketing   = _f(row[10])
        infra       = _f(row[11])
        tecnologia  = _f(row[12])
        frota       = _f(row[13])
        desp_admin  = _f(row[14])
        impostos    = _f(row[15])
        irpj        = _f(row[16])
        desp_op  = atendimento + pessoal + marketing + infra + tecnologia + frota + desp_admin
        encargos = enc_trabal + impostos + irpj
        conn.execute(
            """INSERT OR REPLACE INTO GC_DFC_Mensal
               (AnoMes, Ano, Mes, Entradas, CMV, DespOp, Encargos, DespFin, Outros,
                TotalSaidas, SaldoPeriodo, SaldoAcumulado,
                Pessoal, EncargosTrabalh, Marketing_DFC, Infraestrutura, Tecnologia,
                Frota, DespAdmin, Atendimento, Impostos, IRPJCSLL)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (row[2], row[0], row[1],
             _f(row[3]), _f(row[4]), desp_op, encargos, _f(row[5]), _f(row[6]),
             _f(row[17]), _f(row[18]), _f(row[19]),
             pessoal, enc_trabal, marketing, infra, tecnologia,
             frota, desp_admin, atendimento, impostos, irpj)
        )

    # --- CAC Mensal ---
    ws = wb['📈 CAC Mensal']
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not isinstance(row[0], int):
            continue
        conn.execute(
            "INSERT OR REPLACE INTO GC_CAC_Mensal VALUES (?,?,?,?,?,?,?,?,?,?)",
            (row[2], row[0], row[1],
             _f(row[3]), _f(row[4]), _f(row[5]), _f(row[6]),
             _f(row[7]), int(row[8] or 0), _f(row[9]))
        )

    # --- Lançamentos (v9: cols 18=CategoriaEstruturada, 19=CAPEX/OPEX) ---
    ws = wb['📊 Dados para DRE']
    batch = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not isinstance(row[0], int):
            continue
        batch.append((
            row[0], str(row[1]), row[2],
            row[3], row[4], row[5], row[6], row[7], row[8],
            row[9], _dt(row[10]), _dt(row[11]), _dt(row[12]),
            _f(row[13]), str(row[14]) if row[14] else None,
            row[15], row[16], row[17],
            str(row[18]) if len(row) > 18 and row[18] else None,
            str(row[19]) if len(row) > 19 and row[19] else None,
        ))
    conn.executemany("""
        INSERT INTO GC_Lancamentos
        (Ano, Mes, AnoMes, GrupoDRE, SubgrupoDRE, PlanoContas, CentroCusto, Fornecedor, CNPJ,
         Situacao, DataCompetencia, DataVencimento, DataConfirmacao, Valor, NFe,
         CodLancamento, Loja, Descricao, CategoriaEstruturada, CapexOpex)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, batch)

    # --- DRE Estruturado (fonte direta para api_dre2_dre_anual) ---
    _LABEL_MAP = {
        'RECEITA BRUTA':               'receita_bruta',
        'Impostos sobre Vendas':       'impostos_vendas',
        'RECEITA LÍQUIDA':             'receita_liq',
        'Compras / Materiais':         'cmv',
        'LUCRO BRUTO':                 'lucro_bruto',
        'Pessoal + Pró-labore':        'pessoal',
        'Encargos Trabalhistas':       'enc_trabalh',
        'Marketing e Publicidade':     'marketing',
        'Infraestrutura':              'infraestrutura',
        'Tecnologia e Conectividade':  'tecnologia',
        'Frota e Combustível':         'frota',
        'Atendimento ao Cliente':      'atendimento',
        'Demais Despesas Administrativas': 'desp_admin',
        'EBITDA':                      'ebitda',
        'EBIT':                        'ebit',
        'Despesas Financeiras':        'desp_fin',
        'Outros / Extraordinários':    'outros',
        'LUCRO ANTES DO IR':           'resultado',
        'IRPJ / CSLL':                 'irpj_csll',
        'LUCRO LÍQUIDO':               'lucro_liq',
    }
    ws_est = wb['📋 DRE Estruturado']
    est_rows = list(ws_est.iter_rows(min_row=1, values_only=True))
    # detecta colunas de ano na linha 3 (index 2)
    year_cols = {}
    if len(est_rows) > 2:
        for ci, hv in enumerate(est_rows[2]):
            if hv:
                try:
                    yr = int(str(hv).strip().split('\n')[0].strip().split('(')[0].strip())
                    if 2000 <= yr <= 2100:
                        year_cols[ci] = yr
                except (ValueError, AttributeError):
                    pass
    for row in est_rows[3:]:
        label = None
        for cell_val in row[:3]:
            if cell_val and str(cell_val).strip():
                label = str(cell_val).strip().lstrip('▶').lstrip('–').lstrip('-').strip()
                break
        if not label or label.startswith('%'):
            continue
        campo = None
        for key, fld in _LABEL_MAP.items():
            if key.lower() in label.lower():
                campo = fld
                break
        if not campo:
            continue
        for ci, yr in year_cols.items():
            raw = row[ci] if ci < len(row) else None
            val_f = None if raw is None else _f(raw)
            conn.execute(
                "INSERT OR REPLACE INTO GC_DRE_Estruturado (Ano, Campo, Valor) VALUES (?,?,?)",
                (yr, campo, val_f)
            )

    # --- CAPEX vs OPEX ---
    ws_co = wb['📊 CAPEX vs OPEX']
    co_rows = list(ws_co.iter_rows(min_row=1, values_only=True))
    # detecta linha de cabeçalho de anos (procura a linha com pelo menos 2 anos válidos)
    year_cols_co = {}
    header_row_idx = 0
    for ri, row in enumerate(co_rows):
        found = {}
        for ci, hv in enumerate(row):
            if hv:
                try:
                    yr = int(str(hv).strip().split('\n')[0].strip().split('(')[0].strip())
                    if 2000 <= yr <= 2100:
                        found[ci] = yr
                except (ValueError, AttributeError):
                    pass
        if len(found) >= 2:
            year_cols_co = found
            header_row_idx = ri
            break
    current_secao = 'OPEX'
    ordem_co      = 0
    rb_count      = 0
    for row in co_rows[header_row_idx + 1:]:
        label = None
        for cv in row[:2]:
            if cv and str(cv).strip():
                label = str(cv).strip()
                break
        if not label:
            continue
        clean = label.lstrip('►').lstrip('–').lstrip('-').strip()
        cu = clean.upper()
        # detecta seção
        if 'CAPEX' in cu and 'TOTAL' not in cu and 'OPEX' not in cu:
            current_secao = 'CAPEX'; continue
        if 'OPEX' in cu and 'TOTAL' not in cu and 'CAPEX' not in cu:
            current_secao = 'OPEX'; continue
        if 'TOTAL GERAL' in cu:
            current_secao = 'TOTAL_GERAL'; rb_count = 0; continue
        # determina secao da linha
        if 'TOTAL CAPEX' in cu:
            secao = 'TOTAL_CAPEX'
        elif 'TOTAL OPEX' in cu:
            secao = 'TOTAL_OPEX'
        elif current_secao == 'TOTAL_GERAL':
            secao = 'RB_LANCAMENTOS' if rb_count == 0 else 'RB_DRE'
            rb_count += 1
        else:
            secao = current_secao
        for ci, yr in year_cols_co.items():
            raw = row[ci] if ci < len(row) else None
            conn.execute(
                "INSERT OR REPLACE INTO GC_CAPEX_OPEX_Sheet (Secao, Categoria, Ordem, Ano, Valor) VALUES (?,?,?,?,?)",
                (secao, clean, ordem_co, yr, None if raw is None else _f(raw))
            )
        ordem_co += 1

    conn.commit()
    wb.close()

    return {
        'dre':         conn.execute("SELECT COUNT(*) FROM GC_DRE_Completo").fetchone()[0],
        'dfc':         conn.execute("SELECT COUNT(*) FROM GC_DFC_Mensal").fetchone()[0],
        'cac':         conn.execute("SELECT COUNT(*) FROM GC_CAC_Mensal").fetchone()[0],
        'lancamentos': conn.execute("SELECT COUNT(*) FROM GC_Lancamentos").fetchone()[0],
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@dre2_bp.route('/api/dre2/importar', methods=['POST'])
@login_required
def api_dre2_importar():
    if current_user.username != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    file_bytes = request.files['file'].read()
    conn = get_db()
    try:
        _ensure_tables(conn)
        counts = _import_excel(conn, file_bytes)
        logger.info("GestaoCompleta importada: %s", counts)
        return jsonify({'ok': True, 'counts': counts})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@dre2_bp.route('/api/dre2/status')
@login_required
def api_dre2_status():
    conn = get_db()
    try:
        _ensure_tables(conn)
        rows = conn.execute("SELECT COUNT(*) FROM GC_DRE_Completo").fetchone()[0]
        return jsonify({'imported': rows > 0, 'rows': rows})
    except Exception:
        return jsonify({'imported': False, 'rows': 0})
    finally:
        conn.close()


@dre2_bp.route('/api/dre2/anos')
@login_required
def api_dre2_anos():
    conn = get_db()
    try:
        _ensure_tables(conn)
        anos = [r[0] for r in conn.execute(
            "SELECT DISTINCT Ano FROM GC_DRE_Completo ORDER BY Ano"
        ).fetchall()]
        return jsonify({'anos': anos})
    except Exception:
        return jsonify({'anos': []})
    finally:
        conn.close()


def _mes_conds(date_from, date_to):
    """Constrói cláusulas WHERE + params para filtro de período nas tabelas mensais.
    date_from / date_to: YYYY-MM-DD — compara apenas os 7 primeiros chars (YYYY-MM) com AnoMes."""
    conds, params = [], []
    if date_from:
        conds.append("AnoMes >= ?"); params.append(date_from[:7])
    if date_to:
        conds.append("AnoMes <= ?"); params.append(date_to[:7])
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    return where, params


@dre2_bp.route('/api/dre2/dre')
@login_required
def api_dre2_dre():
    conn = get_db()
    try:
        _ensure_tables(conn)
        date_from = request.args.get('date_from', '')
        date_to   = request.args.get('date_to',   '')
        where, params = _mes_conds(date_from, date_to)

        rows = conn.execute(f"""
            SELECT AnoMes, Ano, Mes, ReceitaBruta, Recebido, AReceber,
                   CMV, DespOp, Encargos, DespFin, Outros, TotalDespesas, Resultado, Margem
            FROM GC_DRE_Completo {where}
            ORDER BY AnoMes
        """, params).fetchall()

        data = [dict(r) for r in rows]
        r_total   = sum(r['ReceitaBruta']  or 0 for r in data)
        d_total   = sum(r['TotalDespesas'] or 0 for r in data)
        res_total = sum(r['Resultado']     or 0 for r in data)

        return jsonify({
            'data': data,
            'kpis': {
                'receita':   r_total,
                'despesas':  d_total,
                'resultado': res_total,
                'margem':    res_total / r_total if r_total else 0,
            }
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@dre2_bp.route('/api/dre2/dfc')
@login_required
def api_dre2_dfc():
    conn = get_db()
    try:
        _ensure_tables(conn)
        date_from = request.args.get('date_from', '')
        date_to   = request.args.get('date_to',   '')
        where, params = _mes_conds(date_from, date_to)

        rows = conn.execute(f"""
            SELECT AnoMes, Ano, Mes, Entradas, CMV, DespOp, Encargos, DespFin, Outros,
                   TotalSaidas, SaldoPeriodo, SaldoAcumulado,
                   COALESCE(Pessoal,0) AS Pessoal, COALESCE(EncargosTrabalh,0) AS EncargosTrabalh,
                   COALESCE(Marketing_DFC,0) AS Marketing_DFC, COALESCE(Infraestrutura,0) AS Infraestrutura,
                   COALESCE(Tecnologia,0) AS Tecnologia, COALESCE(Frota,0) AS Frota,
                   COALESCE(DespAdmin,0) AS DespAdmin, COALESCE(Atendimento,0) AS Atendimento,
                   COALESCE(Impostos,0) AS Impostos, COALESCE(IRPJCSLL,0) AS IRPJCSLL
            FROM GC_DFC_Mensal {where}
            ORDER BY AnoMes
        """, params).fetchall()

        data = [dict(r) for r in rows]
        ent  = sum(r['Entradas']    or 0 for r in data)
        sai  = sum(r['TotalSaidas'] or 0 for r in data)
        acum = data[-1]['SaldoAcumulado'] if data else 0

        return jsonify({
            'data': data,
            'kpis': {
                'entradas':        ent,
                'saidas':          sai,
                'saldo_periodo':   ent - sai,
                'saldo_acumulado': acum,
            }
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@dre2_bp.route('/api/dre2/cac')
@login_required
def api_dre2_cac():
    conn = get_db()
    try:
        _ensure_tables(conn)
        date_from = request.args.get('date_from', '')
        date_to   = request.args.get('date_to',   '')
        where, params = _mes_conds(date_from, date_to)

        rows = conn.execute(f"""
            SELECT AnoMes, Ano, Mes, Comissionamento, Marketing, MaterialCampo, ONTs,
                   TotalCAC, NInstalacoes, CACUnitario
            FROM GC_CAC_Mensal {where}
            ORDER BY AnoMes
        """, params).fetchall()

        data      = [dict(r) for r in rows]
        cac_total = sum(r['TotalCAC']     or 0 for r in data)
        inst      = sum(r['NInstalacoes'] or 0 for r in data)

        return jsonify({
            'data': data,
            'kpis': {
                'cac_total':   cac_total,
                'instalacoes': inst,
                'cac_medio':   cac_total / inst if inst else 0,
            }
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@dre2_bp.route('/api/dre2/lancamentos')
@login_required
def api_dre2_lancamentos():
    conn = get_db()
    try:
        _ensure_tables(conn)
        ano        = request.args.get('ano',        '')
        mes        = request.args.get('mes',        '')
        grupo      = request.args.get('grupo',      '')
        situacao   = request.args.get('situacao',   '')
        capex_opex = request.args.get('capex_opex', '')
        search     = request.args.get('search',     '')
        date_from  = request.args.get('date_from',  '')
        date_to    = request.args.get('date_to',    '')
        page       = max(1, int(request.args.get('page', 1)))
        per_page   = 50

        conds, params = [], []
        if ano:        conds.append("Ano = ?");                               params.append(int(ano))
        if mes:        conds.append("CAST(SUBSTR(AnoMes,6,2) AS INTEGER)=?"); params.append(int(mes))
        if grupo:      conds.append("GrupoDRE = ?");                          params.append(grupo)
        if situacao:   conds.append("Situacao = ?");                          params.append(situacao)
        if capex_opex: conds.append("CapexOpex = ?");                         params.append(capex_opex)
        if date_from:  conds.append("DataCompetencia >= ?");                  params.append(date_from)
        if date_to:    conds.append("DataCompetencia <= ?");                  params.append(date_to)
        if search:
            conds.append("(Fornecedor LIKE ? OR Descricao LIKE ? OR PlanoContas LIKE ?)")
            params += [f'%{search}%', f'%{search}%', f'%{search}%']

        where = ("WHERE " + " AND ".join(conds)) if conds else ""

        total = conn.execute(f"SELECT COUNT(*) FROM GC_Lancamentos {where}", params).fetchone()[0]
        rows  = conn.execute(f"""
            SELECT ID, AnoMes, GrupoDRE, SubgrupoDRE, PlanoContas, Fornecedor,
                   Situacao, DataCompetencia, Valor, CentroCusto, Loja, Descricao,
                   COALESCE(CategoriaEstruturada,'') AS CategoriaEstruturada,
                   COALESCE(CapexOpex,'') AS CapexOpex
            FROM GC_Lancamentos {where}
            ORDER BY DataCompetencia DESC, ID DESC
            LIMIT ? OFFSET ?
        """, params + [per_page, (page - 1) * per_page]).fetchall()

        grupos   = [r[0] for r in conn.execute(
            "SELECT DISTINCT GrupoDRE FROM GC_Lancamentos WHERE GrupoDRE IS NOT NULL ORDER BY GrupoDRE"
        ).fetchall()]
        situacoes = [r[0] for r in conn.execute(
            "SELECT DISTINCT Situacao FROM GC_Lancamentos WHERE Situacao IS NOT NULL ORDER BY Situacao"
        ).fetchall()]
        capex_opts = [r[0] for r in conn.execute(
            "SELECT DISTINCT CapexOpex FROM GC_Lancamentos WHERE CapexOpex IS NOT NULL AND CapexOpex != '' ORDER BY CapexOpex"
        ).fetchall()]
        anos     = [r[0] for r in conn.execute(
            "SELECT DISTINCT Ano FROM GC_Lancamentos ORDER BY Ano"
        ).fetchall()]

        return jsonify({
            'data':     [dict(r) for r in rows],
            'total':    total,
            'page':     page,
            'per_page': per_page,
            'filters':  {'grupos': grupos, 'situacoes': situacoes, 'anos': anos, 'capex_opts': capex_opts},
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# DRE Estruturado Anual
# ---------------------------------------------------------------------------
@dre2_bp.route('/api/dre2/dre_anual')
@login_required
def api_dre2_dre_anual():
    conn = get_db()
    try:
        _ensure_tables(conn)
        # Lê diretamente da tabela importada da aba "📋 DRE Estruturado" do Excel
        est_rows = conn.execute(
            "SELECT Ano, Campo, Valor FROM GC_DRE_Estruturado ORDER BY Ano"
        ).fetchall()

        # pivot: {ano: {campo: valor}}
        est = {}
        for r in est_rows:
            est.setdefault(r['Ano'], {})[r['Campo']] = r['Valor']

        def g(d, k): return d.get(k) or 0

        years = []
        for ano in sorted(est.keys()):
            d   = est[ano]
            rb  = g(d, 'receita_bruta')
            iss = g(d, 'impostos_vendas')
            rl  = g(d, 'receita_liq')
            cmv = g(d, 'cmv')
            lb  = g(d, 'lucro_bruto')
            pessoal  = g(d, 'pessoal')
            enc_trab = g(d, 'enc_trabalh')
            mkt      = g(d, 'marketing')
            infra    = g(d, 'infraestrutura')
            tec      = g(d, 'tecnologia')
            frota    = g(d, 'frota')
            atend    = g(d, 'atendimento')
            desp_adm = g(d, 'desp_admin')
            do_sum   = pessoal + enc_trab + mkt + infra + tec + frota + atend + desp_adm
            ebitda   = g(d, 'ebitda')
            df       = g(d, 'desp_fin')
            ot       = g(d, 'outros')
            ir       = g(d, 'irpj_csll')
            lair     = g(d, 'resultado')
            ll       = g(d, 'lucro_liq')
            def pct(n, base=rb): return round(n / base * 100, 1) if base else 0
            years.append({
                'ano': ano,
                'receita_bruta':  round(rb,  2), 'impostos_vendas': round(iss, 2),
                'receita_liq':    round(rl,  2), 'pct_rl':  pct(rl),
                'cmv':            round(cmv, 2),
                'lucro_bruto':    round(lb,  2), 'pct_lb':  pct(lb),
                'pessoal':        round(pessoal,  2),
                'enc_trabalh':    round(enc_trab, 2),
                'marketing':      round(mkt,      2),
                'infraestrutura': round(infra,    2),
                'tecnologia':     round(tec,      2),
                'frota':          round(frota,    2),
                'atendimento':    round(atend,    2),
                'desp_admin':     round(desp_adm, 2),
                'desp_op':        round(do_sum,   2),
                'ebitda':         round(ebitda, 2), 'pct_ebitda': pct(ebitda),
                'desp_fin':       round(df,  2),
                'outros':         round(ot,  2),
                'resultado':      round(lair, 2), 'pct_res':  pct(lair),
                'irpj_csll':      round(ir,  2),
                'lucro_liq':      round(ll,  2), 'pct_ll':   pct(ll),
            })
        return jsonify({'anos': years})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# DFC Anual
# ---------------------------------------------------------------------------
@dre2_bp.route('/api/dre2/dfc_anual')
@login_required
def api_dre2_dfc_anual():
    conn = get_db()
    try:
        _ensure_tables(conn)
        rows = conn.execute("""
            SELECT Ano,
                   SUM(Entradas)                    AS entradas,
                   SUM(COALESCE(CMV,0))             AS cmv,
                   SUM(COALESCE(Pessoal,0))         AS pessoal,
                   SUM(COALESCE(EncargosTrabalh,0))  AS enc_trabalh,
                   SUM(COALESCE(Marketing_DFC,0))    AS marketing,
                   SUM(COALESCE(Infraestrutura,0))   AS infraestrutura,
                   SUM(COALESCE(Tecnologia,0))       AS tecnologia,
                   SUM(COALESCE(Frota,0))            AS frota,
                   SUM(COALESCE(DespAdmin,0))        AS desp_admin,
                   SUM(COALESCE(Atendimento,0))      AS atendimento,
                   SUM(COALESCE(Impostos,0))         AS impostos,
                   SUM(COALESCE(IRPJCSLL,0))         AS irpj_csll,
                   SUM(COALESCE(DespFin,0))          AS desp_fin,
                   SUM(COALESCE(Outros,0))           AS outros,
                   SUM(TotalSaidas)                  AS total_saidas,
                   SUM(SaldoPeriodo)                 AS saldo_periodo,
                   MAX(SaldoAcumulado)               AS saldo_acumulado
            FROM GC_DFC_Mensal
            GROUP BY Ano ORDER BY Ano
        """).fetchall()
        return jsonify({'anos': [dict(r) for r in rows]})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# CAPEX vs OPEX Anual
# ---------------------------------------------------------------------------
@dre2_bp.route('/api/dre2/capex_opex')
@login_required
def api_dre2_capex_opex():
    conn = get_db()
    try:
        _ensure_tables(conn)
        sheet_rows = conn.execute("""
            SELECT Secao, Categoria, Ordem, Ano, Valor
            FROM GC_CAPEX_OPEX_Sheet
            ORDER BY Secao, Ordem, Ano
        """).fetchall()

        anos_set = sorted(set(r['Ano'] for r in sheet_rows))

        # Denominador = TOTAL GERAL (CAPEX + OPEX) — igual ao que o Excel usa em "% DESPESAS"
        denom = sum(
            r['Valor'] or 0 for r in sheet_rows
            if r['Secao'] in ('TOTAL_CAPEX', 'TOTAL_OPEX')
        )

        def _pivot(secao):
            cats_ord = {}
            for r in sheet_rows:
                if r['Secao'] != secao: continue
                cat = r['Categoria']
                if cat not in cats_ord:
                    cats_ord[cat] = {'cat': cat, 'ordem': r['Ordem'], 'vals': {}}
                cats_ord[cat]['vals'][r['Ano']] = round(r['Valor'] or 0, 2)
            items = sorted(cats_ord.values(), key=lambda x: x['ordem'])
            for item in items:
                tot = sum(item['vals'].values())
                item['total'] = round(tot, 2)
                item['pct']   = round(tot / denom * 100, 1) if denom else 0
            return items

        def _total_row(secao):
            vals = {}
            for r in sheet_rows:
                if r['Secao'] != secao: continue
                vals[r['Ano']] = round(r['Valor'] or 0, 2)
            tot = sum(vals.values())
            return {'vals': vals, 'total': round(tot, 2),
                    'pct': round(tot / denom * 100, 1) if denom else 0}

        tc = _total_row('TOTAL_CAPEX')
        to = _total_row('TOTAL_OPEX')
        tg_vals = {a: round((tc['vals'].get(a, 0) or 0) + (to['vals'].get(a, 0) or 0), 2)
                   for a in anos_set}
        tg_tot = round(sum(tg_vals.values()), 2)

        return jsonify({
            'anos':           anos_set,
            'capex':          _pivot('CAPEX'),
            'opex':           _pivot('OPEX'),
            'total_capex':    tc,
            'total_opex':     to,
            'total_geral':    {'vals': tg_vals, 'total': tg_tot, 'pct': 100.0 if denom else 0},
            'rb_lancamentos': _total_row('RB_LANCAMENTOS'),
            'rb_dre':         _total_row('RB_DRE'),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

